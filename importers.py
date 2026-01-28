"""ARNBEL + ANKET Entegrasyon - Excel Import Islemleri"""
import openpyxl
from datetime import datetime
from models import db, Mudurluk, Personel
from utils import calculate_triage


def import_kunye_excel(file_path):
    """
    Personel kunye (demografik) verilerini Excel'den import et.

    Beklenen kolonlar:
        TC No, Ad Soyad, Yas, Cinsiyet, Egitim, Medeni Durum,
        Cocuk Sayisi, Kidem (Yil), Mudurluk, Gorev

    Args:
        file_path: Excel dosyasinin yolu

    Returns:
        dict: {'success': bool, 'imported': int, 'errors': list}
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Skip header row
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    imported = 0
    errors = []

    for idx, row in enumerate(rows, start=2):
        try:
            if len(row) < 10:
                continue

            tc_no, ad_soyad, yas, cinsiyet, egitim, medeni, cocuk, kidem, mudurluk_adi, gorev = row[:10]

            if not ad_soyad:
                continue

            # Find or create mudurluk
            mudurluk = Mudurluk.query.filter_by(ad=mudurluk_adi).first()
            if not mudurluk and mudurluk_adi:
                mudurluk = Mudurluk(ad=mudurluk_adi)
                db.session.add(mudurluk)
                db.session.flush()

            mudurluk_id = mudurluk.id if mudurluk else None

            # Find existing or create new personnel
            existing = Personel.query.filter_by(tc_no=tc_no).first() if tc_no else None

            if existing:
                # Update existing
                existing.ad_soyad = ad_soyad
                existing.yas = int(yas) if yas else None
                existing.cinsiyet = cinsiyet
                existing.egitim = egitim
                existing.medeni_durum = medeni
                existing.cocuk_sayisi = int(cocuk) if cocuk else None
                existing.kidem_yil = int(kidem) if kidem else None
                existing.mudurluk_id = mudurluk_id
                existing.gorev = gorev
                existing.updated_at = datetime.utcnow()
            else:
                # Create new
                personel = Personel(
                    tc_no=tc_no,
                    ad_soyad=ad_soyad,
                    yas=int(yas) if yas else None,
                    cinsiyet=cinsiyet,
                    egitim=egitim,
                    medeni_durum=medeni,
                    cocuk_sayisi=int(cocuk) if cocuk else None,
                    kidem_yil=int(kidem) if kidem else None,
                    mudurluk_id=mudurluk_id,
                    gorev=gorev
                )
                db.session.add(personel)

            imported += 1

        except Exception as e:
            errors.append(f"Satir {idx}: {str(e)}")

    db.session.commit()

    return {
        'success': True,
        'imported': imported,
        'errors': errors
    }


def import_mmpi_excel(file_path):
    """
    MMPI test sonuclarini Excel'den import et.

    Beklenen kolonlar:
        TC No, L, F, K, PA, SC, Analiz Metni

    Args:
        file_path: Excel dosyasinin yolu

    Returns:
        dict: {'success': bool, 'imported': int, 'errors': list}
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    imported = 0
    errors = []
    updated_ids = []

    for idx, row in enumerate(rows, start=2):
        try:
            if len(row) < 6:
                continue

            tc_no = row[0]
            l_score = row[1] if len(row) > 1 else None
            f_score = row[2] if len(row) > 2 else None
            k_score = row[3] if len(row) > 3 else None
            pa_score = row[4] if len(row) > 4 else None
            sc_score = row[5] if len(row) > 5 else None
            analiz = row[6] if len(row) > 6 else None

            if not tc_no:
                continue

            personel = Personel.query.filter_by(tc_no=str(tc_no)).first()
            if not personel:
                errors.append(f"Satir {idx}: TC No {tc_no} sistemde bulunamadi")
                continue

            personel.mmpi_l_skoru = float(l_score) if l_score else None
            personel.mmpi_f_skoru = float(f_score) if f_score else None
            personel.mmpi_k_skoru = float(k_score) if k_score else None
            personel.mmpi_pa_skoru = float(pa_score) if pa_score else None
            personel.mmpi_sc_skoru = float(sc_score) if sc_score else None
            personel.mmpi_analiz_metni = analiz
            personel.mmpi_tarih = datetime.now().date()
            personel.updated_at = datetime.utcnow()

            updated_ids.append(personel.id)
            imported += 1

        except Exception as e:
            errors.append(f"Satir {idx}: {str(e)}")

    db.session.commit()

    # Recalculate triage for updated personnel
    for personel_id in updated_ids:
        calculate_triage(personel_id)

    return {
        'success': True,
        'imported': imported,
        'errors': errors
    }


def import_anket_excel(file_path):
    """
    Anket sonuclarini Excel'den import et.

    Beklenen kolonlar:
        TC No, Ruh Sagligi Risk (0-4), Umutsuzluk, Uyku, Istah, Isteksizlik

    Args:
        file_path: Excel dosyasinin yolu

    Returns:
        dict: {'success': bool, 'imported': int, 'errors': list}
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    imported = 0
    errors = []
    updated_ids = []

    for idx, row in enumerate(rows, start=2):
        try:
            if len(row) < 6:
                continue

            tc_no = row[0]
            risk = row[1] if len(row) > 1 else None
            umutsuzluk = row[2] if len(row) > 2 else None
            uyku = row[3] if len(row) > 3 else None
            istah = row[4] if len(row) > 4 else None
            isteksizlik = row[5] if len(row) > 5 else None

            if not tc_no:
                continue

            personel = Personel.query.filter_by(tc_no=str(tc_no)).first()
            if not personel:
                errors.append(f"Satir {idx}: TC No {tc_no} sistemde bulunamadi")
                continue

            personel.anket_ruh_sagligi_risk = int(risk) if risk is not None else 0
            personel.anket_umutsuzluk = to_bool(umutsuzluk)
            personel.anket_uyku_bozuklugu = to_bool(uyku)
            personel.anket_istah_degisimi = to_bool(istah)
            personel.anket_isteksizlik = to_bool(isteksizlik)
            personel.updated_at = datetime.utcnow()

            updated_ids.append(personel.id)
            imported += 1

        except Exception as e:
            errors.append(f"Satir {idx}: {str(e)}")

    db.session.commit()

    # Recalculate triage for updated personnel
    for personel_id in updated_ids:
        calculate_triage(personel_id)

    return {
        'success': True,
        'imported': imported,
        'errors': errors
    }


def to_bool(val):
    """Degeri boolean'a cevir."""
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.lower() in ['evet', 'yes', 'var', '1', 'true', 'e', 'y']
    return bool(val) if val else False


def create_sample_mudurlukler():
    """Ornek mudurlukler olustur (bos veritabani icin)."""
    sample_mudurlukler = [
        ('Zabita Mudurlugu', ''),
        ('Fen Isleri Mudurlugu', ''),
        ('Mali Hizmetler Mudurlugu', ''),
        ('Insan Kaynaklari Mudurlugu', ''),
        ('Bilgi Islem Mudurlugu', ''),
        ('Afet Isleri Mudurlugu', ''),
        ('Temizlik Isleri Mudurlugu', ''),
        ('Park ve Bahceler Mudurlugu', ''),
        ('Kultur ve Sosyal Isler Mudurlugu', ''),
        ('Yazi Isleri Mudurlugu', ''),
    ]

    created = 0
    for ad, emoji in sample_mudurlukler:
        existing = Mudurluk.query.filter_by(ad=ad).first()
        if not existing:
            mudurluk = Mudurluk(ad=ad, emoji=emoji)
            db.session.add(mudurluk)
            created += 1

    db.session.commit()
    return created
